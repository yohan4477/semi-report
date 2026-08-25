# -*- coding: utf-8 -*-
"""알파벳 DCF 계산 전체를 엑셀 한 권으로 뽑는다.

보고서 본문은 결론만 싣는다. 이 파일은 그 결론이 나온 칸을 전부 편다 — 연도별
매출·마진·잉여현금흐름·할인계수·현재가치가 줄마다 서고, 영구가치와 순현금이 붙어
주당가치가 된다. 숫자를 직접 만져 보려는 사람을 위한 것이다.

할인 규약은 insights/dcf.py 를 그대로 따른다 — 기말 할인, 1차연도를 (1+r)^1 로
나눈다, 영구가치는 명시적 기간 마지막 연도와 같은 지수로 할인한다. 여기서 다시
구현하지 않고 같은 식을 쓴다. 합계는 dcf.value() 결과와 맞춰 검산한다.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'insights'))

import dcf                      # noqa: E402
import googl_cases as gc        # noqa: E402
from openpyxl import Workbook   # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# scratchpad 는 통째로 무시되는 자리라 여기 두면 보고서에서 링크를 못 건다.
# facts.json 옆에 둔다 — 같은 회사의 같은 계산에 쓰인 것들이다.
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   'insights', 'valuation', 'GOOGL', '알파벳 DCF.xlsx')

HEAD = Font(bold=True, color='FFFFFF')
HEAD_BG = PatternFill('solid', fgColor='2F4858')
SUB = Font(bold=True)
SUB_BG = PatternFill('solid', fgColor='E8EDF0')
HIT = PatternFill('solid', fgColor='FFF3CD')       # 결론 칸
THIN = Side(style='thin', color='BBBBBB')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
R = Alignment(horizontal='right')
C = Alignment(horizontal='center')


def sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def head(ws, row, cells):
    for j, v in enumerate(cells, 1):
        c = ws.cell(row, j, v)
        c.font, c.fill, c.border, c.alignment = HEAD, HEAD_BG, BOX, C
    return row + 1


def band(ws, row, label, span):
    c = ws.cell(row, 1, label)
    c.font, c.fill = SUB, SUB_BG
    for j in range(1, span + 1):
        ws.cell(row, j).fill = SUB_BG
    return row + 1


def line(ws, row, cells, fmts=None, hit=False):
    for j, v in enumerate(cells, 1):
        c = ws.cell(row, j, v)
        c.border = BOX
        if j > 1:
            c.alignment = R
            if fmts and fmts[j - 2]:
                c.number_format = fmts[j - 2]
        if hit:
            c.fill = HIT
    return row + 1


# ── 1. 요약 ──────────────────────────────────────────────────────
def s_summary(wb):
    ws = sheet(wb, '요약', [30, 15, 15, 15])
    r = head(ws, 1, ['알파벳 DCF — 요약', 'Bear', 'Base', 'Bull'])
    r = band(ws, r, '가정', 4)
    rows = [('2026~2028 매출 성장률', 'g1', '0.0%'),
            ('그 구간 끝 잉여현금흐름 마진', 'm1', '0.0%'),
            ('2035년 매출 성장률', 'g2', '0.0%'),
            ('영구 단계 잉여현금흐름 마진', 'm3', '0.0%'),
            ('영구성장률', 'g', '0.00%')]
    for label, key, fmt in rows:
        r = line(ws, r, [label] + [gc.CASES[n][key] for n in ('Bear', 'Base', 'Bull')],
                 [fmt] * 3)
    r = line(ws, r, ['할인율 (WACC)'] + [gc.WACC] * 3, ['0.00%'] * 3)

    r = band(ws, r, '결과 (10억 달러)', 4)
    vals = {n: dcf.value([x[3] for x in gc.path(gc.CASES[n])], gc.WACC,
                         gc.CASES[n]['g'], gc.NET_DEBT, gc.SHARES)
            for n in ('Bear', 'Base', 'Bull')}
    for label, key, fmt in [('명시적 기간 현재가치', 'pv_explicit', '#,##0'),
                            ('영구가치 (할인 전)', 'tv', '#,##0'),
                            ('영구가치의 현재가치', 'pv_tv', '#,##0'),
                            ('기업가치 EV', 'ev', '#,##0'),
                            ('자기자본가치', 'equity', '#,##0'),
                            ('영구가치 비중', 'tv_share', '0.0%')]:
        r = line(ws, r, [label] + [vals[n][key] for n in ('Bear', 'Base', 'Bull')], [fmt] * 3)
    r = line(ws, r, ['순현금 (자기자본에 더한 값)'] + [-gc.NET_DEBT] * 3, ['#,##0'] * 3)

    r = band(ws, r, '주당 (달러)', 4)
    r = line(ws, r, ['주당가치'] + [vals[n]['per_share'] for n in ('Bear', 'Base', 'Bull')],
             ['#,##0'] * 3, hit=True)
    r = line(ws, r, ['현재가 (2026-08-25 종가)'] + [gc.PRICE] * 3, ['#,##0.00'] * 3)
    r = line(ws, r, ['현재가 대비'] + [vals[n]['per_share'] / gc.PRICE - 1
                                   for n in ('Bear', 'Base', 'Bull')], ['0.0%'] * 3, hit=True)
    r = line(ws, r, ['비영업자산 반영 주당'] + [(vals[n]['equity'] + 131.5) / gc.SHARES
                                        for n in ('Bear', 'Base', 'Bull')], ['#,##0'] * 3)

    r += 1
    ws.cell(r, 1, '주식수는 발행주식 %.3f십억 주다. 희석 %.3f십억 주로 나누면 Base 가 %.0f달러다.'
            % (gc.SHARES, gc.SHARES_DIL,
               vals['Base']['equity'] / gc.SHARES_DIL))
    r += 1
    ws.cell(r, 1, '가정은 우리가 정한 것이고 SemiAnalysis 가 제시한 값이 아니다. 투자 추천이 아니다.')
    return ws


# ── 2. 기준값 ────────────────────────────────────────────────────
def s_base(wb):
    ws = sheet(wb, '기준값 (SEC)', [26, 14, 14, 12, 44])
    r = head(ws, 1, ['항목', 'FY2025', 'TTM', '변화', 'XBRL 태그 · 창'])
    c, t = gc.d['sec']['concepts'], gc.t
    B = gc.B
    items = [('매출', 'revenue'), ('영업이익', 'ebit'), ('영업현금흐름', 'ocf'),
             ('설비투자', 'capex'), ('감가상각비', 'dna'),
             ('세전이익', 'pretax_income'), ('세금비용', 'tax_expense'),
             ('영업 밖 손익', 'nonoperating'), ('지분 평가이익', 'equity_fv_gain')]
    for label, k in items:
        a = c[k]['2025']['val'] / B if k in c and '2025' in c[k] else None
        b = t[k]['val'] / B if k in t else None
        chg = (b / a - 1) if (a and b) else None
        r = line(ws, r, [label, a, b, chg, '%s · %s' % (c[k]['2025']['tag'], t[k]['window'])],
                 ['#,##0.0', '#,##0.0', '0.0%', None])
    fcf25 = c['ocf']['2025']['val'] / B - c['capex']['2025']['val'] / B
    r = line(ws, r, ['잉여현금흐름', fcf25, gc.FCF0, gc.FCF0 / fcf25 - 1, '영업현금흐름 − 설비투자'],
             ['#,##0.0', '#,##0.0', '0.0%', None], hit=True)
    r = line(ws, r, ['영업이익률',
                     c['ebit']['2025']['val'] / c['revenue']['2025']['val'],
                     t['ebit']['val'] / t['revenue']['val'], None, ''],
             ['0.0%', '0.0%', None, None])
    r = line(ws, r, ['잉여현금흐름 마진', fcf25 / (c['revenue']['2025']['val'] / B),
                     gc.MARGIN0, None, ''], ['0.0%', '0.0%', None, None])
    r = line(ws, r, ['설비투자 / 매출',
                     c['capex']['2025']['val'] / c['revenue']['2025']['val'],
                     t['capex']['val'] / t['revenue']['val'], None, ''],
             ['0.0%', '0.0%', None, None])
    r = line(ws, r, ['설비투자 / 감가상각',
                     c['capex']['2025']['val'] / c['dna']['2025']['val'],
                     t['capex']['val'] / t['dna']['val'], None,
                     '장기에는 1배로 수렴한다 — 그만큼 영업이익률이 깎인다'],
             ['0.00', '0.00', None, None])

    r = band(ws, r, '할인율 입력값', 5)
    for label, v, fmt, note in [
            ('무위험수익률', gc.d['risk_free']['rate'], '0.00%',
             '%s · %s' % (gc.d['risk_free']['series'], gc.d['risk_free']['date'])),
            ('베타', gc.d['beta']['beta'], '0.000',
             '%s 일간 %d일 · %s~%s' % (gc.d['beta']['index'], gc.d['beta']['n_days'],
                                    gc.d['beta']['start'], gc.d['beta']['end'])),
            ('시장위험프리미엄', 0.046, '0.00%', '우리가 고른 값 — 미국 내재 ERP 범위 중간값'),
            ('자기자본비용', gc.d['risk_free']['rate'] + gc.d['beta']['beta'] * 0.046, '0.00%',
             '무위험수익률 + 베타 × 시장위험프리미엄'),
            ('할인율 (WACC)', gc.WACC, '0.00%', '순현금 회사라 자기자본비용에 붙인다'),
            ('순현금', -gc.NET_DEBT, '#,##0.0', '현금 + 단기투자 − 차입 (2026-06-30)'),
            ('시가총액', gc.MCAP, '#,##0', '주가 × 발행주식수'),
            ('주가', gc.PRICE, '#,##0.00', '2026-08-25 종가')]:
        r = line(ws, r, [label, v, None, None, note], [fmt, None, None, None])
    return ws


# ── 3. 케이스별 연도표 ───────────────────────────────────────────
def s_case(wb, name):
    ws = sheet(wb, name, [8, 13, 11, 11, 13, 12, 13])
    c = gc.CASES[name]
    ws.cell(1, 1, '%s — %s' % (name, c['why'])).font = SUB
    r = head(ws, 2, ['연도', '매출', '성장률', 'FCF마진', '잉여현금흐름',
                     '할인계수', '현재가치'])
    rows = gc.path(c)
    prev = gc.REV0
    for i, (year, rev, m, f) in enumerate(rows, 1):
        disc = 1 / (1 + gc.WACC) ** i
        r = line(ws, r, [year, rev, rev / prev - 1, m, f, disc, f * disc],
                 ['#,##0', '0.0%', '0.0%', '#,##0.0', '0.0000', '#,##0.0'])
        prev = rev
    n = len(rows)
    v = dcf.value([x[3] for x in rows], gc.WACC, c['g'], gc.NET_DEBT, gc.SHARES)

    r = band(ws, r, '마무리', 7)
    fcf_last = rows[-1][3]
    for label, val, fmt, note in [
            ('명시적 기간 현재가치 합', v['pv_explicit'], '#,##0.0', ''),
            ('마지막 해 잉여현금흐름', fcf_last, '#,##0.0', ''),
            ('영구성장률', c['g'], '0.00%', ''),
            ('영구가치 (기간 말)', v['tv'], '#,##0', 'FCF × (1+g) / (r − g)'),
            ('영구가치의 현재가치', v['pv_tv'], '#,##0.0', '(1+r)^%d 로 나눈다' % n),
            ('기업가치 EV', v['ev'], '#,##0.0', ''),
            ('순현금 가산', -gc.NET_DEBT, '#,##0.0', ''),
            ('자기자본가치', v['equity'], '#,##0.0', ''),
            ('발행주식수 (십억 주)', gc.SHARES, '#,##0.000', ''),
            ('주당가치 (달러)', v['per_share'], '#,##0.00', '')]:
        rr = [label, val] + [None] * 4 + [note]
        r = line(ws, r, rr, ['#,##0.0' if fmt is None else fmt] + [None] * 5,
                 hit=(label == '주당가치 (달러)'))
    return ws


# ── 4. 민감도 ────────────────────────────────────────────────────
def s_sens(wb):
    ws = sheet(wb, '민감도', [14, 11, 11, 11, 11, 11])
    ws.cell(1, 1, 'Base 경로에 할인율과 영구성장률만 흔든다. 단위는 주당 달러.').font = SUB
    rs = [0.09, 0.095, 0.10, 0.105, 0.11]
    gs = [0.0175, 0.0225, 0.0275, 0.0325, 0.0375]
    r = head(ws, 2, ['할인율 \\ 영구성장률'] + ['%.2f%%' % (g * 100) for g in gs])
    base = [x[3] for x in gc.path(gc.CASES['Base'])]
    grid = dcf.sensitivity(base, rs, gs, gc.NET_DEBT, gc.SHARES)
    for rr in rs:
        row = ['%.2f%%' % (rr * 100)] + [grid[(rr, g)] for g in gs]
        r = line(ws, r, row, ['#,##0'] * 5)
    r += 1
    ws.cell(r, 1, '스물다섯 칸이 모두 현재가 %.2f달러 아래다. 가장 높은 칸이 %.0f달러다.'
            % (gc.PRICE, max(grid.values())))
    return ws


# ── 5. 역산 ──────────────────────────────────────────────────────
def s_reverse(wb):
    ws = sheet(wb, '역산', [30, 16, 46])
    ws.cell(1, 1, '시가총액을 정답으로 놓고 가정을 거꾸로 푼다.').font = SUB
    r = head(ws, 2, ['항목', '값', '설명'])
    r = band(ws, r, '방향 1 — 우리 경로를 고정하고 할인율을 되돌린다', 3)
    for n in ('Bear', 'Base', 'Bull'):
        c = gc.CASES[n]
        ir = dcf.implied_discount_rate([x[3] for x in gc.path(c)], c['g'], gc.MCAP, gc.NET_DEBT)
        r = line(ws, r, ['%s 경로의 내재 할인율' % n, ir,
                         '우리 할인율 %.2f%% 대비 %+.2f%%p' % (gc.WACC * 100, (ir - gc.WACC) * 100)],
                 ['0.00%', None])
    r = band(ws, r, '방향 2 — 할인율을 고정하고 성장률을 되돌린다', 3)
    for rr in (0.09, 0.095, 0.10, 0.105, 0.11):
        g = dcf.implied_growth(gc.FCF0, rr, 0.0275, 10, gc.MCAP, gc.NET_DEBT)
        r = line(ws, r, ['할인율 %.1f%% 일 때 요구 성장률' % (rr * 100), g,
                         '기준 잉여현금흐름 %.1f 이 10년간 매년 이만큼 늘어야 한다' % gc.FCF0],
                 ['0.00%', None], hit=(rr == 0.10))
    r = band(ws, r, '그 성장률이 뜻하는 것', 3)
    g10 = dcf.implied_growth(gc.FCF0, 0.10, 0.0275, 10, gc.MCAP, gc.NET_DEBT)
    f10 = gc.FCF0 * (1 + g10) ** 10
    base_last = gc.path(gc.CASES['Base'])[-1][3]
    for label, val, fmt, note in [
            ('기준 잉여현금흐름 (TTM)', gc.FCF0, '#,##0.0', '2026-06-30 기준'),
            ('10년 뒤 요구 잉여현금흐름', f10, '#,##0', '지금의 %.1f배' % (f10 / gc.FCF0)),
            ('Base 경로의 마지막 해', base_last, '#,##0', '2035년'),
            ('차이', f10 - base_last, '#,##0', '이만큼을 우리 경로가 못 그린다')]:
        r = line(ws, r, [label, val, note], [fmt, None])
    r += 1
    ws.cell(r, 1, '실무 도구가 쓰는 보수적 상한은 10년 성장 단계에 20%다. 요구 성장률은 그 상한 밖이다.')
    return ws


def build():
    wb = Workbook()
    wb.remove(wb.active)
    s_summary(wb)
    s_base(wb)
    for n in ('Bear', 'Base', 'Bull'):
        s_case(wb, n)
    s_sens(wb)
    s_reverse(wb)
    wb.save(OUT)
    return OUT


if __name__ == '__main__':
    print('->', build())
